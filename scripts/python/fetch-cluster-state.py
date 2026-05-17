#!/usr/bin/env python3
"""
fetch-cluster-state.py

Snapshot the current state of every OpenShift project starting with "pid-",
grouped by stage (ref/prod/test/phase/pnext/other), and produce:

  * per-namespace JSON snapshot (normalized current state)
  * per-namespace YAML "desired state" (re-applyable manifests, stripped of
    status/runtime fields) — can be diffed or piped into `oc apply -f`
  * an HPA-binding validation report (does each HPA's scaleTargetRef resolve
    to an existing Deployment/StatefulSet, are min/max replicas sane, are
    metrics configured, ...)
  * CSV summaries with the most important dimensions across namespaces
  * a top-level overview JSON
  * a multi-sheet Excel workbook (_cluster-state.xlsx) with frozen headers,
    autofilters and conditional formatting for HPA issues and quota usage

Layout (under ./state-YYYYMMDD-HHMMSS/):

    by-stage/<stage>/<ns>/
        snapshot.json
        hpa-bindings.json
        desired/
            00-namespace.yaml
            10-resourcequotas.yaml
            20-limitranges.yaml
            30-networkpolicies.yaml
            40-deployments.yaml
            41-statefulsets.yaml
            50-services.yaml
            60-pvcs.yaml
            70-hpas.yaml
    _overview.json
    _hpa-validation.csv
    _dimensions.csv

Stage detection mirrors check-bind-resources.sh:
    pid-<id>-<app>-<STAGE>-<num>-<suffix>

Usage:
    python3 fetch-cluster-state.py [--output-dir DIR]
                                   [--stage STAGE [--stage STAGE ...]]
                                   [--project NS [--project NS ...]]
                                   [--workers N]

Requires:
    * oc (logged in)
    * Python 3.6.8+
    * PyYAML   (pip install pyyaml)
    * openpyxl (pip install openpyxl)   — for the .xlsx workbook
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

try:
    import yaml
except ImportError:
    sys.stderr.write("error: PyYAML is required (pip install pyyaml)\n")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


STAGE_KEYWORDS = ("ref", "prod", "test", "phase", "pnext")
PROJECT_PREFIX = "pid-"

# Allow callers on hosts with a non-standard oc location (e.g. a network-
# mounted /mnt/Gruppenfreigabe/linux-bin/oc) to point the script at it
# directly, bypassing PATH entirely. Falls back to "oc" so existing setups
# keep working unchanged. Picked up once at import time so all subsequent
# subprocess.run([_OC_BIN, ...]) calls use the same binary.
_OC_BIN = os.environ.get("OC_BIN", "oc")


# ---------------------------------------------------------------------------
# oc wrappers
# ---------------------------------------------------------------------------

def oc(*args, check=False):
    """Fuehrt einen `oc`-Befehl aus und gibt dessen stdout als String zurueck.

    Beispiel: oc("get", "pods", "-n", "default") -> Roh-Ausgabe von oc get pods.

    Wenn check=True ist und der Befehl mit Exit != 0 endet, wird eine
    Exception weitergereicht. Mit check=False (Default) wird die Fehlermeldung
    nach stderr geschrieben und ein leerer String zurueckgegeben -- das
    macht das Aufrufen in JSON-Parsern stabiler (leerer String -> '{}').
    """
    try:
        result = subprocess.run(
            [_OC_BIN] + list(args),
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            universal_newlines=True,
            check=check,
        )
        # With check=False, subprocess.run does NOT raise on non-zero exit
        # -- it silently returns the CompletedProcess. Without this block,
        # any oc error (RBAC, missing API resource, wrong namespace) would
        # look exactly like "no resources found": empty stdout, empty list
        # downstream. Surface the stderr so it shows up in run.log.
        if result.returncode != 0:
            err = (result.stderr or "").strip()
            sys.stderr.write(
                "{} {} exited {}{}\n".format(
                    _OC_BIN, " ".join(args), result.returncode,
                    ": " + err if err else "",
                )
            )
        return result.stdout
    except FileNotFoundError:
        sys.stderr.write(
            "error: '{}' is not installed or not on PATH. "
            "Set OC_BIN=/full/path/to/oc to point at it explicitly.\n"
            .format(_OC_BIN))
        sys.exit(1)
    except subprocess.CalledProcessError as e:
        sys.stderr.write(f"{_OC_BIN} {' '.join(args)} failed: {e.stderr}\n")
        if check:
            raise
        return ""


def oc_get_single(kind, name, ns=None):
    """Holt eine einzelne K8s-Ressource als geparstes JSON-Dict.

    Beispiel: oc_get_single("namespace", "default")
    Liefert None, wenn die Ressource nicht existiert oder die Antwort nicht
    parsebar ist -- der Aufrufer kann mit 'if obj is None' einfach pruefen.
    """
    args = ["get", kind, name, "-o", "json"]
    if ns:
        args += ["-n", ns]
    out = oc(*args)
    if not out.strip():
        return None
    try:
        return json.loads(out)
    except json.JSONDecodeError:
        return None


def oc_get_list(kind, ns):
    """Holt eine Liste K8s-Ressourcen aus einem Namespace.

    Gibt die '.items'-Liste zurueck -- also direkt das was man iterieren
    moechte. Bei Fehler/leerer Antwort eine leere Liste, damit
    `for obj in oc_get_list(...)` immer sicher ist.
    """
    out = oc("get", kind, "-n", ns, "-o", "json")
    if not out.strip():
        return []
    try:
        return json.loads(out).get("items", [])
    except json.JSONDecodeError:
        return []


def list_pid_projects():
    """Liefert alle OpenShift-Projekte, deren Name mit 'pid-' beginnt.

    Wird verwendet, wenn der User keine expliziten --project-Flags setzt:
    dann scannt der Snapshotter automatisch alle pid-*-Namespaces.
    Die Liste ist alphabetisch sortiert, damit Ausgaben stabil bleiben.
    """
    out = oc("get", "projects", "-o", "name")
    names = []
    for line in out.splitlines():
        if "/" in line:
            line = line.split("/", 1)[1]
        if line.startswith(PROJECT_PREFIX):
            names.append(line)
    return sorted(names)


# ---------------------------------------------------------------------------
# Stage detection + parsing helpers
# ---------------------------------------------------------------------------

def detect_stage(ns):
    """Bestimmt die Stage (ref/prod/test/phase/pnext/other) aus dem Namespace.

    Konvention: pid-<id>-<app>-<STAGE>-<num>-<suffix>. Zuerst wird Position 3
    geprueft (dort steht die Stage), dann als Fallback jedes weitere Segment.
    Findet sich nichts: 'other'. Identische Logik wie in den bash-Loops.
    """
    parts = ns.lower().split("-")
    if len(parts) > 3 and parts[3] in STAGE_KEYWORDS:
        return parts[3]
    for p in parts:
        if p in STAGE_KEYWORDS:
            return p
    return "other"


def parse_cpu_millis(v):
    """Wandelt eine K8s-CPU-Angabe in Millicores (int) um.

    '500m' -> 500, '0.5' -> 500, '1' -> 1000. Leerer/None/ungueltiger
    Input liefert 0, damit Summen einfach mit '+' gebildet werden koennen.
    """
    if v in (None, "", "<none>"):
        return 0
    s = str(v).strip()
    if s.endswith("m"):
        try:
            return int(float(s[:-1]))
        except ValueError:
            return 0
    try:
        return int(float(s) * 1000)
    except ValueError:
        return 0


_MEM_UNITS = {
    "Ki": 1 / 1024,
    "Mi": 1,
    "Gi": 1024,
    "Ti": 1024 * 1024,
    "K":  1000 / (1024 * 1024),
    "M":  1000 ** 2 / (1024 * 1024),
    "G":  1000 ** 3 / (1024 * 1024),
    "T":  1000 ** 4 / (1024 * 1024),
}


def parse_mem_mib(v):
    """Wandelt eine K8s-Memory-Angabe in MiB (int) um.

    Versteht binaere (Ki/Mi/Gi/Ti) und dezimale (K/M/G/T) Suffixe.
    Ohne Suffix wird der Wert als Bytes interpretiert. Liefert 0 fuer
    leeren/None/ungueltigen Input.
    """
    if v in (None, "", "<none>"):
        return 0
    s = str(v).strip()
    m = re.match(r"^([0-9]+(?:\.[0-9]+)?)([A-Za-z]+)?$", s)
    if not m:
        return 0
    n = float(m.group(1))
    unit = m.group(2) or ""
    if unit == "":
        factor = 1 / (1024 * 1024)
    else:
        factor = _MEM_UNITS.get(unit, 0)
    return int(n * factor)


def parse_storage_gib(v):
    """Wandelt eine PVC-Storage-Angabe in GiB (float, 2 Nachkommastellen) um.

    Delegiert an parse_mem_mib und teilt durch 1024 -- so funktionieren
    alle Suffixe automatisch (z. B. '10Gi', '5000Mi', '1Ti').
    """
    mib = parse_mem_mib(v)
    return round(mib / 1024, 2)


# ---------------------------------------------------------------------------
# Desired-state stripping
# ---------------------------------------------------------------------------

_STRIP_METADATA_FIELDS = {
    "creationTimestamp", "resourceVersion", "uid", "generation",
    "selfLink", "managedFields", "ownerReferences", "finalizers",
}
_STRIP_ANNOTATION_PREFIXES = (
    "kubectl.kubernetes.io/last-applied-configuration",
    "deployment.kubernetes.io/revision",
    "openshift.io/generated-by",
    "kubernetes.io/change-cause",
    "autoscaling.alpha.kubernetes.io/conditions",
    "autoscaling.alpha.kubernetes.io/current-metrics",
)


def _clean_metadata(meta):
    """Entfernt server-injizierte Felder aus einem metadata-Block.

    Felder wie 'resourceVersion', 'uid', 'managedFields' und der
    'last-applied-configuration'-Annotations-Eintrag werden vom Cluster
    erzeugt und sind beim Re-Apply stoerend (oder unmoeglich neu zu
    setzen). Wird beim Schreiben der 'desired/'-YAMLs verwendet.
    """
    if not isinstance(meta, dict):
        return meta
    clean = {k: v for k, v in meta.items() if k not in _STRIP_METADATA_FIELDS}
    anns = clean.get("annotations") or {}
    anns = {
        k: v for k, v in anns.items()
        if not any(k.startswith(p) for p in _STRIP_ANNOTATION_PREFIXES)
    }
    if anns:
        clean["annotations"] = anns
    elif "annotations" in clean:
        clean.pop("annotations")
    return clean


def to_desired(obj):
    """Reduziert ein K8s-Live-Objekt auf eine 'desired'-Form, die wiederapplybar ist.

    Behaelt: apiVersion, kind, bereinigte metadata, spec, data, stringData,
    type. Wirft alles weg, was 'status' / 'runtime' ist (z. B. assigned
    NodePorts, readyReplicas, condition timestamps).

    Ergebnis kann via `oc apply -f` zurueck in den Cluster gepushed werden
    und ist gut zum Diffen gegen Git geeignet.
    """
    if not isinstance(obj, dict):
        return obj
    out = {
        "apiVersion": obj.get("apiVersion"),
        "kind": obj.get("kind"),
        "metadata": _clean_metadata(obj.get("metadata", {})),
    }
    for k in ("spec", "data", "stringData", "type"):
        if k in obj:
            out[k] = obj[k]
    return out


# ---------------------------------------------------------------------------
# Per-workload dimensions
# ---------------------------------------------------------------------------

def workload_dimensions(workload):
    """Extrahiert die wichtigen Kennzahlen aus einem Deployment/StatefulSet.

    Liefert:
      replicas, readyReplicas, availableReplicas, containers (Anzahl), images,
      cpu_req_millis, mem_req_mib, cpu_lim_millis, mem_lim_mib

    Wichtig: CPU/Memory sind PRO POD summiert (alle Container zusammen),
    NICHT pro Pod mal Replicas. Den Cluster-Footprint berechnet erst der
    Aggregator durch Multiplikation mit replicas.
    """
    spec = workload.get("spec") or {}
    tmpl_spec = ((spec.get("template") or {}).get("spec") or {})
    containers = tmpl_spec.get("containers") or []

    cpu_req = mem_req = cpu_lim = mem_lim = 0
    for c in containers:
        resources = c.get("resources") or {}
        req = resources.get("requests") or {}
        lim = resources.get("limits") or {}
        cpu_req += parse_cpu_millis(req.get("cpu"))
        mem_req += parse_mem_mib(req.get("memory"))
        cpu_lim += parse_cpu_millis(lim.get("cpu"))
        mem_lim += parse_mem_mib(lim.get("memory"))

    status = workload.get("status") or {}
    return {
        "replicas": spec.get("replicas"),
        "readyReplicas": status.get("readyReplicas"),
        "availableReplicas": status.get("availableReplicas"),
        "containers": len(containers),
        "images": [c.get("image") for c in containers],
        "cpu_req_millis": cpu_req,
        "mem_req_mib": mem_req,
        "cpu_lim_millis": cpu_lim,
        "mem_lim_mib": mem_lim,
    }


# ---------------------------------------------------------------------------
# HPA binding validation
# ---------------------------------------------------------------------------

def _metric_summary(m):
    """Verdichtet einen HPA-Metric-Eintrag auf {type, name, target}.

    Eine HPA-Spec kann verschiedene Metric-Typen haben ('Resource', 'Pods',
    'Object', 'External', 'ContainerResource'). Diese Funktion vereinheitlicht
    die Struktur, damit der Reporter sie ueberall gleich darstellen kann.
    """
    t = m.get("type")
    if t == "Resource":
        r = m.get("resource") or {}
        return {"type": t, "name": r.get("name"), "target": r.get("target")}
    for key in ("pods", "object", "external", "containerResource"):
        if key in m:
            sub = m.get(key) or {}
            metric = sub.get("metric") or {}
            return {"type": t, "name": metric.get("name") or sub.get("name"),
                    "target": sub.get("target")}
    return {"type": t}


def validate_hpas(hpas, workloads_by_kind):
    """Prueft jeden HPA gegen seinen Ziel-Workload und liefert ein Findings-Dict pro HPA.

    Gefundene Probleme landen in einer 'issues'-Liste pro HPA. Geprueft wird:
      - scaleTargetRef.name nicht leer
      - scaleTargetRef.kind ist Deployment oder StatefulSet
      - referenziertes Workload existiert im Namespace
      - minReplicas / maxReplicas gesetzt und min <= max
      - mindestens eine Metric konfiguriert
      - bei Resource-Metric: Ziel-Container muessen resources.requests haben
        (sonst bleibt der HPA inaktiv)
      - ScalingActive / AbleToScale = False werden als Issues uebernommen

    'ok' im Ergebnis ist True genau dann, wenn issues leer ist.
    """
    results = []
    for hpa in hpas:
        meta = hpa.get("metadata") or {}
        spec = hpa.get("spec") or {}
        status = hpa.get("status") or {}
        ref = spec.get("scaleTargetRef") or {}
        kind = ref.get("kind", "")
        name = ref.get("name", "")
        target = workloads_by_kind.get(kind, {}).get(name)

        issues = []
        if not name:
            issues.append("scaleTargetRef.name is empty")
        if kind not in ("Deployment", "StatefulSet"):
            issues.append(f"unsupported scaleTargetRef.kind={kind!r}")
        if name and target is None:
            issues.append(f"target {kind}/{name} not found in namespace")

        min_r = spec.get("minReplicas")
        max_r = spec.get("maxReplicas")
        if min_r is None:
            issues.append("minReplicas missing")
        if max_r is None:
            issues.append("maxReplicas missing")
        if isinstance(min_r, int) and isinstance(max_r, int) and min_r > max_r:
            issues.append(f"minReplicas ({min_r}) > maxReplicas ({max_r})")

        metrics = spec.get("metrics") or []
        if not metrics:
            issues.append("no metrics configured")

        target_spec_replicas = None
        target_has_resource_requests = None
        if isinstance(target, dict):
            target_spec_replicas = (target.get("spec") or {}).get("replicas")
            target_has_resource_requests = _target_has_requests(target)
            # HPA on Resource metric requires container resources.requests on target.
            if any(m.get("type") == "Resource" for m in metrics) and not target_has_resource_requests:
                issues.append(
                    "HPA uses Resource metric but target containers have no resources.requests"
                )

        cond_unhealthy = [
            c for c in (status.get("conditions") or [])
            if c.get("type") in ("ScalingActive", "AbleToScale")
            and c.get("status") == "False"
        ]
        for c in cond_unhealthy:
            issues.append(f"condition {c.get('type')}=False ({c.get('reason')})")

        results.append({
            "hpa": meta.get("name"),
            "namespace": meta.get("namespace"),
            "targetKind": kind,
            "targetName": name,
            "targetFound": target is not None,
            "targetSpecReplicas": target_spec_replicas,
            "targetHasRequests": target_has_resource_requests,
            "minReplicas": min_r,
            "maxReplicas": max_r,
            "currentReplicas": status.get("currentReplicas"),
            "desiredReplicas": status.get("desiredReplicas"),
            "metrics": [_metric_summary(m) for m in metrics],
            "conditions": [
                {"type": c.get("type"), "status": c.get("status"),
                 "reason": c.get("reason"), "message": c.get("message")}
                for c in (status.get("conditions") or [])
            ],
            "issues": issues,
            "ok": len(issues) == 0,
        })
    return results


def _target_has_requests(target):
    """True, wenn ALLE Container des Targets resources.requests gesetzt haben.

    Wichtig fuer HPA-Validierung: ein Resource-Metric-HPA braucht
    requests auf jedem Container, sonst kann er die Utilization nicht
    berechnen und bleibt inaktiv.
    """
    containers = (((target.get("spec") or {}).get("template") or {})
                  .get("spec") or {}).get("containers") or []
    if not containers:
        return False
    return all((c.get("resources") or {}).get("requests") for c in containers)


# ---------------------------------------------------------------------------
# Per-namespace pipeline
# ---------------------------------------------------------------------------

def fetch_namespace(ns):
    """Holt alle relevanten K8s-Ressourcen eines Namespaces in einem Aufruf.

    Liefert ein Dict mit Listen der Ressourcen-Typen (deployments, statefulsets,
    hpas, services, pvcs, resourcequotas, limitranges, networkpolicies) plus
    dem Namespace-Objekt selbst. Wird einmal pro Namespace aufgerufen und
    dient als Eingabe fuer alle weiteren Per-Namespace-Verarbeitungsschritte.
    """
    return {
        "namespace_obj":   oc_get_single("namespace", ns),
        "deployments":     oc_get_list("deployments", ns),
        "statefulsets":    oc_get_list("statefulsets", ns),
        "hpas":            oc_get_list("hpa", ns),
        "services":        oc_get_list("services", ns),
        "pvcs":            oc_get_list("pvc", ns),
        "resourcequotas":  oc_get_list("resourcequota", ns),
        "limitranges":     oc_get_list("limitrange", ns),
        "networkpolicies": oc_get_list("networkpolicy", ns),
    }


def build_snapshot(ns, stage, raw, bindings):
    """Verdichtet die Roh-K8s-Objekte zu einem schlanken snapshot.json-Dict.

    Aus dem (potentiell sehr grossen) Rohformat von 'oc get -o json' wird
    eine flache Darstellung gebaut: pro Ressource nur die Felder, die in
    Reports und Aggregationen tatsaechlich gebraucht werden. Wird sowohl
    als per-Namespace snapshot.json gespeichert als auch in das Excel-
    Workbook eingespeist.
    """
    ns_obj = raw["namespace_obj"] or {}
    labels = ((ns_obj.get("metadata") or {}).get("labels")) or {}

    return {
        "namespace": ns,
        "stage": stage,
        "env": labels.get("environment"),
        "labels": labels,
        "deployments": [
            {"name": d["metadata"]["name"], **workload_dimensions(d)}
            for d in raw["deployments"]
        ],
        "statefulsets": [
            {"name": s["metadata"]["name"], **workload_dimensions(s)}
            for s in raw["statefulsets"]
        ],
        "hpas": [
            {
                "name": b["hpa"], "target": f"{b['targetKind']}/{b['targetName']}",
                "found": b["targetFound"],
                "min": b["minReplicas"], "max": b["maxReplicas"],
                "current": b["currentReplicas"], "desired": b["desiredReplicas"],
                "ok": b["ok"], "issues": b["issues"],
            }
            for b in bindings
        ],
        "services": [
            {
                "name": s["metadata"]["name"],
                "type": (s.get("spec") or {}).get("type"),
                "clusterIP": (s.get("spec") or {}).get("clusterIP"),
                "ports": (s.get("spec") or {}).get("ports") or [],
            }
            for s in raw["services"]
        ],
        "pvcs": [
            {
                "name": p["metadata"]["name"],
                "status": (p.get("status") or {}).get("phase"),
                "storage_gib": parse_storage_gib(
                    ((p.get("status") or {}).get("capacity") or {}).get("storage")
                    or ((p.get("spec") or {}).get("resources") or {}).get("requests", {}).get("storage")
                ),
                "storageClass": (p.get("spec") or {}).get("storageClassName"),
                "accessModes": (p.get("spec") or {}).get("accessModes") or [],
            }
            for p in raw["pvcs"]
        ],
        "resourceQuotas": [
            {"name": q["metadata"]["name"],
             "hard": (q.get("spec") or {}).get("hard"),
             "used": (q.get("status") or {}).get("used")}
            for q in raw["resourcequotas"]
        ],
        "limitRanges": [
            {"name": lr["metadata"]["name"],
             "limits": (lr.get("spec") or {}).get("limits")}
            for lr in raw["limitranges"]
        ],
        "networkPolicies": [np["metadata"]["name"] for np in raw["networkpolicies"]],
    }


def write_desired(desired_dir, raw):
    """Schreibt 'desired/'-YAML-Files (wiederapplybar) pro Ressourcen-Typ.

    Dateinamen sind numerisch praefixiert (00, 10, 20, ...), damit
    `oc apply -f desired/` sie in der richtigen Reihenfolge anwendet
    (Namespace zuerst, danach Quota/LimitRange/NetPol, danach Workloads,
    am Ende HPAs). Leere Listen werden uebersprungen, also keine
    leeren YAML-Files.
    """
    desired_dir.mkdir(parents=True, exist_ok=True)
    files = [
        ("00-namespace.yaml",       [raw["namespace_obj"]] if raw["namespace_obj"] else []),
        ("10-resourcequotas.yaml",  raw["resourcequotas"]),
        ("20-limitranges.yaml",     raw["limitranges"]),
        ("30-networkpolicies.yaml", raw["networkpolicies"]),
        ("40-deployments.yaml",     raw["deployments"]),
        ("41-statefulsets.yaml",    raw["statefulsets"]),
        ("50-services.yaml",        raw["services"]),
        ("60-pvcs.yaml",            raw["pvcs"]),
        ("70-hpas.yaml",            raw["hpas"]),
    ]
    for filename, items in files:
        items = [it for it in items if it]
        if not items:
            continue
        with (desired_dir / filename).open("w") as f:
            yaml.safe_dump_all((to_desired(it) for it in items), f,
                               sort_keys=False)


def process_namespace(ns, stage, out_root):
    """Komplette Per-Namespace-Pipeline: fetch -> validate -> snapshot -> write.

    Schritte:
      1. fetch_namespace() holt alle Ressourcen
      2. validate_hpas() prueft jeden HPA -> hpa-bindings.json
      3. build_snapshot() macht die schlanke Darstellung -> snapshot.json
      4. write_desired() schreibt die wiederapplybaren YAMLs

    Wird (im Parallel-Modus) gleichzeitig fuer mehrere Namespaces aufgerufen.
    Liefert ein Aggregate-Dict, das main() spaeter fuer cluster-weite CSVs
    und den Excel-Report nutzt.
    """
    ns_dir = out_root / "by-stage" / stage / ns
    ns_dir.mkdir(parents=True, exist_ok=True)

    raw = fetch_namespace(ns)

    workloads_by_kind = {
        "Deployment":  {d["metadata"]["name"]: d for d in raw["deployments"]},
        "StatefulSet": {s["metadata"]["name"]: s for s in raw["statefulsets"]},
    }
    bindings = validate_hpas(raw["hpas"], workloads_by_kind)
    snapshot = build_snapshot(ns, stage, raw, bindings)

    (ns_dir / "snapshot.json").write_text(json.dumps(snapshot, indent=2, default=str))
    (ns_dir / "hpa-bindings.json").write_text(json.dumps(bindings, indent=2, default=str))
    write_desired(ns_dir / "desired", raw)

    return {
        "namespace": ns,
        "stage": stage,
        "deployments": raw["deployments"],
        "statefulsets": raw["statefulsets"],
        "bindings": bindings,
        "snapshot": snapshot,
    }


# ---------------------------------------------------------------------------
# CSV writers
# ---------------------------------------------------------------------------

def write_hpa_csv(path, rows):
    """Schreibt die aggregierte _hpa-validation.csv (eine Zeile pro HPA, clusterweit).

    Spalten-Reihenfolge ist fest -- so bleiben Filter und Pivots in
    Excel stabil. 'issues' ist semikolongetrennt, damit man in der
    Tabellen-Software danach filtern kann ('contains: not found').
    """
    fields = [
        "stage", "namespace", "hpa", "ok",
        "targetKind", "targetName", "targetFound", "targetSpecReplicas",
        "targetHasRequests",
        "minReplicas", "maxReplicas", "currentReplicas", "desiredReplicas",
        "metricsCount", "issues",
    ]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_dimensions_csv(path, rows):
    """Schreibt die aggregierte _dimensions.csv (eine Zeile pro Workload).

    Pro Deployment/StatefulSet eine Zeile mit den von workload_dimensions()
    extrahierten Kennzahlen plus Stage/Namespace/Kind/Name. Eignet sich
    fuer Capacity-Reviews und Sizing-Pivots.
    """
    fields = [
        "stage", "namespace", "kind", "name",
        "replicas", "readyReplicas", "containers", "images",
        "cpu_req_millis", "mem_req_mib", "cpu_lim_millis", "mem_lim_mib",
    ]
    with path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ---------------------------------------------------------------------------
# Excel workbook
# ---------------------------------------------------------------------------

_HEADER_FILL = "1F4E78"      # dark blue
_HEADER_FONT = "FFFFFF"      # white
_ROW_BAD     = "F8CBAD"      # light red   — HPA issues, quota >95%
_ROW_WARN    = "FFE699"      # light amber — quota 80–95%
_ROW_OK      = "E2EFDA"      # light green — clean rows


def _parse_quota_amount(dim, val):
    """Normalisiert einen ResourceQuota-Wert auf eine vergleichbare Zahl.

    Eine Quota hat unterschiedliche Dimensionen mit unterschiedlichen
    Einheiten:
      - 'requests.cpu' / 'limits.cpu'       -> Millicores
      - 'requests.memory' / 'storage'       -> MiB
      - 'pods', 'services', 'count/*', ...  -> Stueckzahl (float)

    Anhand des Dimensions-Namens wird die richtige Einheit gewaehlt --
    so kann 'used / hard' als Prozentsatz berechnet werden.
    """
    if val in (None, ""):
        return 0.0
    d = dim.lower()
    if "cpu" in d:
        return float(parse_cpu_millis(val))
    if "memory" in d or "storage" in d:
        return float(parse_mem_mib(val))
    try:
        return float(str(val))
    except ValueError:
        return 0.0


def _style_header(ws, ncols):
    """Stylt die erste Zeile eines Excel-Sheets als Header.

    Blauer Hintergrund + weiss + bold, plus 'freeze pane' (Header bleibt
    beim Scrollen sichtbar) und ein Auto-Filter auf der ganzen
    Header-Zeile. Wird auf jedes Sheet im Workbook angewendet.
    """
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT)
    align = Alignment(vertical="center")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"


def _autosize(ws, max_width=60):
    """Setzt die Spaltenbreite jedes Excel-Sheets auf den laengsten Inhalt.

    Begrenzt auf max_width (Default 60 Zeichen), damit eine lange
    Issue-Liste die Tabelle nicht ungenutzt breit macht. Mindestbreite
    10, damit die Header lesbar bleiben.
    """
    for col_cells in ws.columns:
        letter = col_cells[0].column_letter
        width = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            for line in str(v).split("\n"):
                if len(line) > width:
                    width = len(line)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def _fill_row(ws, row_idx, ncols, color):
    """Faerbt eine komplette Zeile in einem Excel-Sheet ein.

    Wird fuer 'conditional formatting' verwendet: rote Zeile bei HPA-Issue
    oder PVC unbound, Amber bei Quota 80-95%, Gruen bei sauberen HPAs.
    """
    fill = PatternFill("solid", fgColor=color)
    for c in range(1, ncols + 1):
        ws.cell(row=row_idx, column=c).fill = fill


def _write_rows(ws, headers, rows, row_colors=None):
    """Schreibt Header + Datenzeilen in ein Excel-Sheet und stylt es.

    Wird von jeder _build_*_sheet()-Funktion verwendet, damit das Layout
    (Header-Styling, Auto-Size, optionale Zeilenfarbe) ueberall gleich
    ist. row_colors ist eine Liste in derselben Reihenfolge wie rows;
    None pro Zeile bedeutet 'nicht einfaerben'.
    """
    ws.append(headers)
    for i, row in enumerate(rows):
        ws.append(row)
        if row_colors and row_colors[i]:
            _fill_row(ws, ws.max_row, len(headers), row_colors[i])
    _style_header(ws, len(headers))
    _autosize(ws)


def _format_ports(ports):
    """Formatiert eine Service-ports-Liste als kompakten String.

    Beispiele:
      [{port:80, protocol:'TCP'}]                   -> '80/TCP'
      [{name:'http', port:80, targetPort:8080}]     -> 'http:80/TCP->8080'

    Wird im Excel-Services-Sheet verwendet, damit alle Ports in einer
    Zelle stehen.
    """
    out = []
    for p in ports or []:
        name = p.get("name") or ""
        port = p.get("port")
        proto = p.get("protocol") or "TCP"
        target = p.get("targetPort")
        chunk = f"{name}:{port}/{proto}" if name else f"{port}/{proto}"
        if target not in (None, "", port):
            chunk += f"→{target}"
        out.append(chunk)
    return ", ".join(out)


def _build_overview_sheet(wb, overview):
    """Baut das erste Sheet 'Overview' im Excel-Workbook.

    Enthaelt Metadaten (Generierungszeit, Cluster-URL, Namespace-Anzahl)
    und eine Per-Stage-Zusammenfassung (namespaces, deployments,
    statefulsets, hpas, hpaIssues, pvcs). Rot eingefaerbt fuer Stages
    mit HPA-Issues.
    """
    ws = wb.create_sheet("Overview")
    ws["A1"] = "Cluster forensic snapshot"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Generated at"
    ws["B2"] = overview.get("generatedAt")
    ws["A3"] = "Cluster"
    ws["B3"] = overview.get("cluster")
    ws["A4"] = "Total namespaces"
    ws["B4"] = overview.get("totalNamespaces")
    for r in range(2, 5):
        ws.cell(row=r, column=1).font = Font(bold=True)

    ws["A6"] = "Per-stage summary"
    ws["A6"].font = Font(bold=True, size=12)

    headers = ["stage", "namespaces", "deployments", "statefulsets",
               "hpas", "hpaIssues", "pvcs"]
    ws.append([])  # row 7 spacer? actually append goes to next free row
    # Re-anchor: clear append cursor by writing header at row 8
    for i, h in enumerate(headers, start=1):
        ws.cell(row=8, column=i, value=h)
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    font = Font(bold=True, color=_HEADER_FONT)
    for i in range(1, len(headers) + 1):
        c = ws.cell(row=8, column=i)
        c.fill = fill
        c.font = font

    row = 9
    for stage in sorted((overview.get("byStage") or {}).keys()):
        s = overview["byStage"][stage]
        ws.cell(row=row, column=1, value=stage)
        ws.cell(row=row, column=2, value=s.get("namespaces"))
        ws.cell(row=row, column=3, value=s.get("deployments"))
        ws.cell(row=row, column=4, value=s.get("statefulsets"))
        ws.cell(row=row, column=5, value=s.get("hpas"))
        ws.cell(row=row, column=6, value=s.get("hpaIssues"))
        ws.cell(row=row, column=7, value=s.get("pvcs"))
        if (s.get("hpaIssues") or 0) > 0:
            _fill_row(ws, row, len(headers), _ROW_BAD)
        row += 1
    _autosize(ws)


def _build_namespaces_sheet(wb, results):
    """Sheet 'Namespaces': eine Zeile pro Namespace mit Counts pro Ressource.

    Spalten: stage, namespace, env, deployments, statefulsets, hpas,
    hpaIssues, pvcs, services, resourceQuotas, limitRanges,
    networkPolicies. Rot, wenn der Namespace HPA-Issues hat.
    """
    ws = wb.create_sheet("Namespaces")
    headers = ["stage", "namespace", "env",
               "deployments", "statefulsets", "hpas", "hpaIssues",
               "pvcs", "services", "resourceQuotas", "limitRanges",
               "networkPolicies"]
    rows, colors = [], []
    for res in sorted(results, key=lambda r: (r["stage"], r["namespace"])):
        snap = res["snapshot"]
        bindings = res["bindings"]
        hpa_issues = sum(1 for b in bindings if not b["ok"])
        rows.append([
            res["stage"], res["namespace"], snap.get("env"),
            len(snap["deployments"]), len(snap["statefulsets"]),
            len(bindings), hpa_issues,
            len(snap["pvcs"]), len(snap["services"]),
            len(snap["resourceQuotas"]), len(snap["limitRanges"]),
            len(snap["networkPolicies"]),
        ])
        colors.append(_ROW_BAD if hpa_issues else None)
    _write_rows(ws, headers, rows, colors)


def _build_hpa_sheet(wb, results):
    """Sheet 'HPAs': eine Zeile pro HorizontalPodAutoscaler.

    Die fehlerhaften HPAs werden NACH OBEN sortiert (False vor True),
    damit man die Probleme zuerst sieht. Rot bei issues, Gruen bei
    sauberen HPAs.
    """
    ws = wb.create_sheet("HPAs")
    headers = ["stage", "namespace", "hpa", "ok",
               "targetKind", "targetName", "targetFound",
               "targetSpecReplicas", "targetHasRequests",
               "minReplicas", "maxReplicas",
               "currentReplicas", "desiredReplicas",
               "metricsCount", "metrics", "issues"]
    rows, colors = [], []
    for res in results:
        for b in res["bindings"]:
            metrics_str = "; ".join(
                f"{m.get('type')}:{m.get('name') or ''}" for m in b["metrics"]
            )
            rows.append([
                res["stage"], res["namespace"], b["hpa"], b["ok"],
                b["targetKind"], b["targetName"], b["targetFound"],
                b["targetSpecReplicas"], b["targetHasRequests"],
                b["minReplicas"], b["maxReplicas"],
                b["currentReplicas"], b["desiredReplicas"],
                len(b["metrics"]), metrics_str,
                "; ".join(b["issues"]),
            ])
            colors.append(_ROW_BAD if not b["ok"] else _ROW_OK)
    rows.sort(key=lambda r: (bool(r[3]), r[0], r[1], r[2]))  # bad (False) first
    colors = [_ROW_BAD if not r[3] else _ROW_OK for r in rows]
    _write_rows(ws, headers, rows, colors)


def _build_quotas_sheet(wb, results):
    """Sheet 'ResourceQuotas': eine Zeile pro Quota-Dimension.

    used%-Spalte zeigt die Auslastung. Rot bei >=95%, Amber bei >=80%,
    sonst neutral. Hilft beim Aufspueren von Namespaces, die kurz vorm
    Quota-Limit stehen.
    """
    ws = wb.create_sheet("ResourceQuotas")
    headers = ["stage", "namespace", "quota", "dimension",
               "hard", "used", "used %"]
    rows, colors = [], []
    for res in results:
        for q in res["snapshot"]["resourceQuotas"]:
            hard = q.get("hard") or {}
            used = q.get("used") or {}
            dims = sorted(set(hard.keys()) | set(used.keys()))
            for dim in dims:
                h_raw = hard.get(dim, "")
                u_raw = used.get(dim, "")
                h = _parse_quota_amount(dim, h_raw)
                u = _parse_quota_amount(dim, u_raw)
                pct = round((u / h) * 100, 1) if h > 0 else None
                rows.append([
                    res["stage"], res["namespace"], q["name"], dim,
                    str(h_raw) if h_raw != "" else "",
                    str(u_raw) if u_raw != "" else "",
                    pct,
                ])
                if pct is None:
                    colors.append(None)
                elif pct >= 95:
                    colors.append(_ROW_BAD)
                elif pct >= 80:
                    colors.append(_ROW_WARN)
                else:
                    colors.append(None)
    _write_rows(ws, headers, rows, colors)


def _build_limitranges_sheet(wb, results):
    """Sheet 'LimitRanges': eine Zeile pro LimitRange-Constraint x Ressource.

    Zeigt die Default/Min/Max-Werte fuer cpu, memory, ... pro LimitRange.
    Hilfreich um zu pruefen, ob ein Namespace ueberhaupt Default-Limits
    setzt.
    """
    ws = wb.create_sheet("LimitRanges")
    headers = ["stage", "namespace", "limitRange", "type", "resource",
               "min", "max", "default", "defaultRequest",
               "maxLimitRequestRatio"]
    rows = []
    for res in results:
        for lr in res["snapshot"]["limitRanges"]:
            name = lr["name"]
            for limit in lr.get("limits") or []:
                ltype = limit.get("type")
                resources = set()
                for key in ("max", "min", "default",
                            "defaultRequest", "maxLimitRequestRatio"):
                    resources.update((limit.get(key) or {}).keys())
                for r in sorted(resources):
                    rows.append([
                        res["stage"], res["namespace"], name, ltype, r,
                        (limit.get("min") or {}).get(r, ""),
                        (limit.get("max") or {}).get(r, ""),
                        (limit.get("default") or {}).get(r, ""),
                        (limit.get("defaultRequest") or {}).get(r, ""),
                        (limit.get("maxLimitRequestRatio") or {}).get(r, ""),
                    ])
    _write_rows(ws, headers, rows)


def _build_workloads_sheet(wb, results):
    """Sheet 'Workloads': eine Zeile pro Deployment/StatefulSet.

    Zeigt Replicas, Ready, Container-Anzahl, Images und CPU/Mem
    Requests+Limits. Amber, wenn ein Workload weder cpu_req noch
    mem_req gesetzt hat (hat Folgen fuer HPAs und Scheduler-
    Reservierungen).
    """
    ws = wb.create_sheet("Workloads")
    headers = ["stage", "namespace", "kind", "name",
               "replicas", "readyReplicas", "availableReplicas",
               "containers", "images",
               "cpu_req_millis", "mem_req_mib",
               "cpu_lim_millis", "mem_lim_mib"]
    rows, colors = [], []
    for res in results:
        for kind, key in (("Deployment", "deployments"),
                          ("StatefulSet", "statefulsets")):
            for w in res["snapshot"][key]:
                rows.append([
                    res["stage"], res["namespace"], kind, w["name"],
                    w.get("replicas"),
                    w.get("readyReplicas"),
                    w.get("availableReplicas"),
                    w.get("containers"),
                    ", ".join(w.get("images") or []),
                    w.get("cpu_req_millis"),
                    w.get("mem_req_mib"),
                    w.get("cpu_lim_millis"),
                    w.get("mem_lim_mib"),
                ])
                # Highlight workloads with no resource requests configured
                no_req = (w.get("cpu_req_millis") or 0) == 0 and (
                    w.get("mem_req_mib") or 0) == 0
                colors.append(_ROW_WARN if no_req else None)
    _write_rows(ws, headers, rows, colors)


def _build_pvcs_sheet(wb, results):
    """Sheet 'PVCs': eine Zeile pro PersistentVolumeClaim.

    Rot, wenn der PVC nicht 'Bound' ist (z. B. 'Pending' wegen falscher
    StorageClass). Storage-Groesse in GiB.
    """
    ws = wb.create_sheet("PVCs")
    headers = ["stage", "namespace", "pvc", "status",
               "storage_gib", "storageClass", "accessModes"]
    rows, colors = [], []
    for res in results:
        for p in res["snapshot"]["pvcs"]:
            rows.append([
                res["stage"], res["namespace"], p["name"],
                p.get("status"),
                p.get("storage_gib"),
                p.get("storageClass"),
                ", ".join(p.get("accessModes") or []),
            ])
            colors.append(_ROW_BAD if p.get("status") not in (None, "Bound")
                          else None)
    _write_rows(ws, headers, rows, colors)


def _build_services_sheet(wb, results):
    """Sheet 'Services': eine Zeile pro Service.

    Zeigt type (ClusterIP/NodePort/LoadBalancer/Headless), clusterIP und
    eine kompakte Ports-Darstellung (via _format_ports).
    """
    ws = wb.create_sheet("Services")
    headers = ["stage", "namespace", "service", "type",
               "clusterIP", "ports"]
    rows = []
    for res in results:
        for s in res["snapshot"]["services"]:
            rows.append([
                res["stage"], res["namespace"], s["name"],
                s.get("type"), s.get("clusterIP"),
                _format_ports(s.get("ports")),
            ])
    _write_rows(ws, headers, rows)


# ---------------------------------------------------------------------------
# Text inspector (--text mode)
#
# Druckt einen einzelnen Namespace menschenlesbar nach stdout. Gedacht
# fuer Debugging: zeigt, wieviel oc tatsaechlich geliefert hat, und ob
# das im Snapshot landet. Wenn die Excel-Sheets oder die CSVs leer sind,
# sieht man hier auf welcher Stufe die Daten verschwinden.
# ---------------------------------------------------------------------------

def _text_hr(char="=", width=78):
    return char * width


def render_namespace_text(res):
    """Druckt einen einzelnen process_namespace()-Result als Text-Report."""
    snap = res.get("snapshot") or {}
    bindings = res.get("bindings") or []
    deployments = res.get("deployments") or []
    statefulsets = res.get("statefulsets") or []

    print()
    print(_text_hr("="))
    print("NAMESPACE  {}   (stage={})".format(res["namespace"], res["stage"]))
    print(_text_hr("="))

    # --- Section 1: raw fetch counts -------------------------------------
    # The first place to look when sheets are empty: how much did oc
    # actually return? If anything here is 0 unexpectedly, the problem
    # is at the oc/fetch layer, not in the renderers.
    counts = [
        ("deployments",     len(deployments)),
        ("statefulsets",    len(statefulsets)),
        ("hpas",            len(bindings)),
        ("services",        len(snap.get("services") or [])),
        ("pvcs",            len(snap.get("pvcs") or [])),
        ("resourceQuotas",  len(snap.get("resourceQuotas") or [])),
        ("limitRanges",     len(snap.get("limitRanges") or [])),
        ("networkPolicies", len(snap.get("networkPolicies") or [])),
    ]
    print()
    print("FETCH COUNTS  (what `oc get <kind> -n {} -o json` returned)".format(res["namespace"]))
    for name, n in counts:
        marker = "  <-- 0; the corresponding sheet/CSV WILL be empty" if n == 0 else ""
        print("  {:<18} {:>4}{}".format(name + ":", n, marker))

    # --- Section 2: per-kind details -------------------------------------
    print()
    print("DEPLOYMENTS ({})".format(len(deployments)))
    if not deployments:
        print("  (none)")
    for d in deployments:
        dims = workload_dimensions(d)
        print("  - {}".format(d["metadata"]["name"]))
        print("      replicas={}  ready={}  containers={}".format(
            dims.get("replicas"), dims.get("readyReplicas"), dims.get("containers")))
        print("      requests cpu={}m mem={}Mi   limits cpu={}m mem={}Mi".format(
            dims.get("cpu_req_millis"), dims.get("mem_req_mib"),
            dims.get("cpu_lim_millis"), dims.get("mem_lim_mib")))
        for img in (dims.get("images") or []):
            print("      image: {}".format(img))

    print()
    print("STATEFULSETS ({})".format(len(statefulsets)))
    if not statefulsets:
        print("  (none)")
    for s in statefulsets:
        dims = workload_dimensions(s)
        print("  - {}  replicas={}  containers={}".format(
            s["metadata"]["name"], dims.get("replicas"), dims.get("containers")))

    print()
    print("HPAs ({})".format(len(bindings)))
    if not bindings:
        print("  (none) -- the _hpa-validation.csv and the HPAs sheet WILL be empty.")
    for b in bindings:
        print("  - {}  ok={}".format(b["hpa"], b["ok"]))
        print("      target:   {}/{}  (found={})".format(
            b["targetKind"], b["targetName"], b["targetFound"]))
        print("      replicas: min={}  max={}  current={}  desired={}".format(
            b["minReplicas"], b["maxReplicas"],
            b["currentReplicas"], b["desiredReplicas"]))
        print("      metrics:  {}".format(len(b.get("metrics") or [])))
        if b.get("issues"):
            print("      issues:")
            for iss in b["issues"]:
                print("        - {}".format(iss))

    print()
    print("SERVICES ({})".format(len(snap.get("services") or [])))
    for svc in snap.get("services") or []:
        ports = ",".join(str(p.get("port")) for p in (svc.get("ports") or []))
        print("  - {}  type={}  ports={}".format(svc["name"], svc.get("type"), ports))

    print()
    print("PVCs ({})".format(len(snap.get("pvcs") or [])))
    for p in snap.get("pvcs") or []:
        print("  - {}  status={}  size={}Gi  class={}".format(
            p["name"], p.get("status"), p.get("storage_gib"), p.get("storageClass")))

    print()
    print("RESOURCEQUOTAS ({})".format(len(snap.get("resourceQuotas") or [])))
    for q in snap.get("resourceQuotas") or []:
        print("  - {}".format(q["name"]))
        for k in sorted((q.get("hard") or {}).keys()):
            hard = (q.get("hard") or {}).get(k)
            used = (q.get("used") or {}).get(k, "?")
            print("      {}: used={} / hard={}".format(k, used, hard))

    print()
    print("LIMITRANGES ({})".format(len(snap.get("limitRanges") or [])))
    for lr in snap.get("limitRanges") or []:
        print("  - {}".format(lr["name"]))
        for limit in (lr.get("limits") or []):
            print("      type={}  default={}  defaultRequest={}".format(
                limit.get("type"), limit.get("default"), limit.get("defaultRequest")))

    print()
    print("NETWORKPOLICIES ({})".format(len(snap.get("networkPolicies") or [])))
    for nm in snap.get("networkPolicies") or []:
        print("  - {}".format(nm))

    # --- Section 3: would-be CSV row counts ------------------------------
    # If these are 0 but FETCH COUNTS above are non-zero, the bug is
    # downstream of the fetch (in validate_hpas/build_snapshot/the
    # renderer). If both are 0, the bug is in the fetch.
    print()
    print("CSV ROW PREVIEW  (what would be written to the aggregate CSVs)")
    print("  _hpa-validation.csv:  {} data row(s)".format(len(bindings)))
    print("  _dimensions.csv:      {} data row(s)  (deployments + statefulsets)".format(
        len(deployments) + len(statefulsets)))
    print()
    if all(n == 0 for _, n in counts):
        print("WARNING: every fetch returned 0 items. Either the namespace is")
        print("genuinely empty, or `oc` is failing silently. Re-run the loop")
        print("after the stderr-surfacing patch and check run.log for lines")
        print("like 'oc get ... exited <N>: ...'.")
    print(_text_hr("="))


def write_excel(path, results, overview):
    """Baut das _cluster-state.xlsx-Workbook mit allen Sheets.

    Sheets: Overview, Namespaces, HPAs, ResourceQuotas, LimitRanges,
    Workloads, PVCs, Services. Jeder Tab hat formatierte Header
    (Auto-Filter, Freeze-Pane) und farbliche Markierung der Probleme.

    Wenn openpyxl nicht installiert ist, wird eine Warnung nach stderr
    geschrieben und das Workbook ausgelassen (kein Abbruch -- CSV und
    JSON-Reports werden trotzdem geschrieben).
    """
    if not _HAS_OPENPYXL:
        sys.stderr.write(
            "warning: openpyxl not installed — skipping .xlsx export "
            "(pip install openpyxl)\n"
        )
        return
    wb = openpyxl.Workbook()
    # Remove default sheet, we'll create our own
    wb.remove(wb.active)
    _build_overview_sheet(wb, overview)
    _build_namespaces_sheet(wb, results)
    _build_hpa_sheet(wb, results)
    _build_quotas_sheet(wb, results)
    _build_limitranges_sheet(wb, results)
    _build_workloads_sheet(wb, results)
    _build_pvcs_sheet(wb, results)
    _build_services_sheet(wb, results)
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    """Einstiegspunkt des Snapshotters.

    Parst CLI-Argumente, ermittelt die zu verarbeitenden Namespaces
    (entweder die per --project gelistet wurden, oder alle pid-*), und
    fuehrt process_namespace() parallel via ThreadPoolExecutor aus.

    Am Ende werden die aggregierten Files (_hpa-validation.csv,
    _dimensions.csv, _overview.json, _cluster-state.xlsx) geschrieben.

    Rueckgabe: 0 bei Erfolg (oder wenn keine Projekte gefunden wurden).
    Nicht-fatale Fehler einzelner Namespaces werden auf stderr gemeldet
    und der Lauf geht weiter.
    """
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--output-dir", default=None,
                    help="Base output directory (default: ./state-<timestamp>)")
    ap.add_argument("--stage", action="append", default=None,
                    help="Limit to one or more stages (repeatable)")
    ap.add_argument("--project", action="append", default=None,
                    help="Limit to specific pid-* projects (repeatable)")
    ap.add_argument("--workers", type=int, default=4,
                    help="Parallel oc workers (default: 4)")
    ap.add_argument("--text", action="store_true",
                    help="Single-namespace debug mode. Requires exactly one "
                         "--project. Prints a plain-text report to stdout "
                         "(fetch counts, per-kind details, would-be CSV row "
                         "counts) so you can see where the data drops out "
                         "when the Excel/CSV come back empty. Writes the "
                         "normal artifacts to a temp dir if --output-dir is "
                         "not given.")
    args = ap.parse_args()

    # --text precondition: one namespace, sequential, no surprises.
    if args.text:
        if not args.project or len(args.project) != 1:
            sys.stderr.write(
                "error: --text requires exactly one --project NS\n")
            return 2
        args.workers = 1

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    if args.text and not args.output_dir:
        import tempfile
        out_root = Path(tempfile.mkdtemp(prefix="fetch-cluster-state-text-"))
        sys.stderr.write(
            "# --text mode: writing artifacts to {} (temp)\n".format(out_root))
    else:
        out_root = Path(args.output_dir or f"./state-{ts}")
    out_root.mkdir(parents=True, exist_ok=True)

    projects = args.project or list_pid_projects()
    if not projects:
        print("No pid-* projects found.")
        return 0

    targets = [(p, detect_stage(p)) for p in projects]
    if args.stage:
        targets = [(p, s) for p, s in targets if s in set(args.stage)]
    if not targets:
        print("No projects match the --stage filter.")
        return 0

    if not args.text:
        print(f"Processing {len(targets)} projects with {args.workers} workers...")

    results = []
    with ThreadPoolExecutor(max_workers=args.workers) as ex:
        future_to_ns = {
            ex.submit(process_namespace, ns, stage, out_root): ns
            for ns, stage in targets
        }
        for fut in as_completed(future_to_ns):
            ns = future_to_ns[fut]
            try:
                res = fut.result()
                if not args.text:
                    print(f"  [ok] {ns}  [stage={res['stage']}]")
                results.append(res)
            except Exception as e:
                print(f"  [ERR] {ns}: {e}", file=sys.stderr)

    # --text mode: print the human-readable report and stop.
    # The CSV/Excel/overview steps below assume "cluster-wide aggregate"
    # output which makes no sense for a single-namespace debug run.
    if args.text:
        if not results:
            sys.stderr.write(
                "error: no result for {} -- see stderr above for the cause.\n"
                .format(args.project[0]))
            return 1
        render_namespace_text(results[0])
        return 0

    # Aggregate CSVs
    hpa_rows = []
    dim_rows = []
    for res in results:
        stage = res["stage"]
        ns = res["namespace"]
        for b in res["bindings"]:
            hpa_rows.append({
                "stage": stage,
                "namespace": ns,
                "hpa": b["hpa"],
                "ok": b["ok"],
                "targetKind": b["targetKind"],
                "targetName": b["targetName"],
                "targetFound": b["targetFound"],
                "targetSpecReplicas": b["targetSpecReplicas"],
                "targetHasRequests": b["targetHasRequests"],
                "minReplicas": b["minReplicas"],
                "maxReplicas": b["maxReplicas"],
                "currentReplicas": b["currentReplicas"],
                "desiredReplicas": b["desiredReplicas"],
                "metricsCount": len(b["metrics"]),
                "issues": "; ".join(b["issues"]),
            })
        for d in res["deployments"]:
            dim = workload_dimensions(d)
            dim_rows.append({
                "stage": stage, "namespace": ns, "kind": "Deployment",
                "name": d["metadata"]["name"],
                "replicas": dim["replicas"],
                "readyReplicas": dim["readyReplicas"],
                "containers": dim["containers"],
                "images": ", ".join(dim["images"] or []),
                "cpu_req_millis": dim["cpu_req_millis"],
                "mem_req_mib": dim["mem_req_mib"],
                "cpu_lim_millis": dim["cpu_lim_millis"],
                "mem_lim_mib": dim["mem_lim_mib"],
            })
        for s in res["statefulsets"]:
            dim = workload_dimensions(s)
            dim_rows.append({
                "stage": stage, "namespace": ns, "kind": "StatefulSet",
                "name": s["metadata"]["name"],
                "replicas": dim["replicas"],
                "readyReplicas": dim["readyReplicas"],
                "containers": dim["containers"],
                "images": ", ".join(dim["images"] or []),
                "cpu_req_millis": dim["cpu_req_millis"],
                "mem_req_mib": dim["mem_req_mib"],
                "cpu_lim_millis": dim["cpu_lim_millis"],
                "mem_lim_mib": dim["mem_lim_mib"],
            })

    write_hpa_csv(out_root / "_hpa-validation.csv", hpa_rows)
    write_dimensions_csv(out_root / "_dimensions.csv", dim_rows)

    overview = {
        "generatedAt": datetime.now().isoformat(),
        "cluster": oc("whoami", "--show-server").strip() or "unknown",
        "totalNamespaces": len(results),
        "byStage": {},
        "namespaces": [],
    }
    for res in results:
        stage = res["stage"]
        snap = res["snapshot"]
        bindings = res["bindings"]
        per_stage = overview["byStage"].setdefault(
            stage, {"namespaces": 0, "deployments": 0, "statefulsets": 0,
                    "hpas": 0, "hpaIssues": 0, "pvcs": 0})
        per_stage["namespaces"] += 1
        per_stage["deployments"] += len(snap["deployments"])
        per_stage["statefulsets"] += len(snap["statefulsets"])
        per_stage["hpas"] += len(bindings)
        per_stage["hpaIssues"] += sum(1 for b in bindings if not b["ok"])
        per_stage["pvcs"] += len(snap["pvcs"])
        overview["namespaces"].append({
            "namespace": res["namespace"],
            "stage": stage,
            "deployments": len(snap["deployments"]),
            "statefulsets": len(snap["statefulsets"]),
            "hpas": len(bindings),
            "hpaIssues": sum(1 for b in bindings if not b["ok"]),
            "pvcs": len(snap["pvcs"]),
            "services": len(snap["services"]),
        })
    overview["namespaces"].sort(key=lambda r: (r["stage"], r["namespace"]))

    (out_root / "_overview.json").write_text(json.dumps(overview, indent=2))

    xlsx_path = out_root / "_cluster-state.xlsx"
    write_excel(xlsx_path, results, overview)

    bad_hpas = sum(1 for r in hpa_rows if not r["ok"])
    print(f"\nDone. Output: {out_root}")
    print(f"  Overview:        {out_root / '_overview.json'}")
    print(f"  HPA validation:  {out_root / '_hpa-validation.csv'}  "
          f"({bad_hpas} HPA(s) with issues)")
    print(f"  Dimensions:      {out_root / '_dimensions.csv'}")
    if _HAS_OPENPYXL:
        print(f"  Excel workbook:  {xlsx_path}")
    print(f"  Per-namespace:   {out_root}/by-stage/<stage>/<ns>/")
    return 0


if __name__ == "__main__":
    sys.exit(main())
